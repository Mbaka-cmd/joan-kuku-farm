# Advanced Reporting & Export System

from django.db import models
from django.utils import timezone
from datetime import datetime, timedelta
import logging

logger = logging.getLogger('reporting')

# ============================================================
# REPORTING MODELS
# ============================================================

class Report(models.Model):
    """Scheduled reports"""
    REPORT_TYPE_CHOICES = [
        ('sales', 'Sales Report'),
        ('inventory', 'Inventory Report'),
        ('customer', 'Customer Report'),
        ('financial', 'Financial Report'),
        ('performance', 'Performance Report'),
        ('custom', 'Custom Report'),
    ]
    
    SCHEDULE_CHOICES = [
        ('daily', 'Daily'),
        ('weekly', 'Weekly'),
        ('monthly', 'Monthly'),
        ('quarterly', 'Quarterly'),
        ('yearly', 'Yearly'),
    ]
    
    FORMAT_CHOICES = [
        ('pdf', 'PDF'),
        ('excel', 'Excel'),
        ('csv', 'CSV'),
        ('json', 'JSON'),
        ('html', 'HTML'),
    ]
    
    # Report info
    name = models.CharField(max_length=255)
    description = models.TextField(blank=True)
    report_type = models.CharField(max_length=30, choices=REPORT_TYPE_CHOICES)
    
    # Configuration
    filters = models.JSONField(default=dict)
    columns = models.JSONField(default=list)
    
    # Schedule
    schedule = models.CharField(max_length=20, choices=SCHEDULE_CHOICES)
    next_run = models.DateTimeField()
    
    # Export
    export_format = models.CharField(max_length=10, choices=FORMAT_CHOICES, default='pdf')
    recipients = models.JSONField(default=list)  # Email addresses
    
    # Owner
    created_by = models.ForeignKey('users.CustomUser', on_delete=models.CASCADE)
    
    # Status
    is_active = models.BooleanField(default=True)
    last_run = models.DateTimeField(null=True, blank=True)
    
    created_at = models.DateTimeField(auto_now_add=True)
    
    class Meta:
        db_table = 'report'
        ordering = ['-created_at']


class ReportExecution(models.Model):
    """Track report executions"""
    STATUS_CHOICES = [
        ('pending', 'Pending'),
        ('generating', 'Generating'),
        ('completed', 'Completed'),
        ('failed', 'Failed'),
    ]
    
    report = models.ForeignKey(Report, on_delete=models.CASCADE)
    
    # Execution
    status = models.CharField(max_length=20, choices=STATUS_CHOICES, default='pending')
    
    # File
    file = models.FileField(upload_to='reports/%Y/%m/%d/', null=True, blank=True)
    
    # Performance
    start_time = models.DateTimeField()
    end_time = models.DateTimeField(null=True, blank=True)
    row_count = models.IntegerField(default=0)
    
    # Error
    error_message = models.TextField(blank=True)
    
    created_at = models.DateTimeField(auto_now_add=True)
    
    class Meta:
        db_table = 'report_execution'
        ordering = ['-created_at']


class ReportTemplate(models.Model):
    """Report templates"""
    name = models.CharField(max_length=255)
    description = models.TextField(blank=True)
    
    # Template
    template_html = models.TextField()
    template_css = models.TextField(blank=True)
    
    # Variables
    variables = models.JSONField(default=list)  # Available variables
    
    # Preview
    preview = models.ImageField(upload_to='report_previews/', null=True, blank=True)
    
    created_at = models.DateTimeField(auto_now_add=True)
    
    class Meta:
        db_table = 'report_template'


class DataExport(models.Model):
    """Track data exports"""
    # Export
    export_id = models.CharField(max_length=50, unique=True)
    
    # Data
    data_type = models.CharField(max_length=100)  # customers, orders, products
    format = models.CharField(max_length=10)  # csv, excel, json
    
    # Filter
    filters = models.JSONField(default=dict)
    
    # File
    file = models.FileField(upload_to='exports/%Y/%m/%d/')
    file_size = models.IntegerField()  # bytes
    
    # Status
    status = models.CharField(
        max_length=20,
        choices=[('preparing', 'Preparing'), ('ready', 'Ready'), ('expired', 'Expired')]
    )
    
    # User
    created_by = models.ForeignKey('users.CustomUser', on_delete=models.CASCADE)
    
    # Expiration
    expires_at = models.DateTimeField()
    
    created_at = models.DateTimeField(auto_now_add=True)
    
    class Meta:
        db_table = 'data_export'
        ordering = ['-created_at']


# ============================================================
# REPORTING ENGINE
# ============================================================

class ReportingEngine:
    """Generate and manage reports"""
    
    @staticmethod
    def generate_sales_report(start_date, end_date, filters=None):
        """Generate sales report"""
        from apps.orders.models import Order, OrderItem
        from django.db.models import Sum, Count, Avg
        
        orders = Order.objects.filter(
            created_at__date__gte=start_date,
            created_at__date__lte=end_date
        )
        
        if filters:
            if 'category' in filters:
                orders = orders.filter(orderitem__product__category=filters['category'])
            if 'status' in filters:
                orders = orders.filter(status=filters['status'])
        
        report_data = {
            'period': f"{start_date} to {end_date}",
            'total_orders': orders.count(),
            'total_revenue': orders.aggregate(Sum('total_amount'))['total_amount__sum'] or 0,
            'avg_order_value': orders.aggregate(Avg('total_amount'))['total_amount__avg'] or 0,
            'total_items': OrderItem.objects.filter(order__in=orders).aggregate(Sum('quantity'))['quantity__sum'] or 0,
        }
        
        # By day
        daily_sales = orders.extra(
            select={'date': 'DATE(created_at)'}
        ).values('date').annotate(
            revenue=Sum('total_amount'),
            orders=Count('id')
        ).order_by('date')
        
        report_data['daily_breakdown'] = list(daily_sales)
        
        return report_data
    
    @staticmethod
    def generate_customer_report(start_date, end_date):
        """Generate customer report"""
        from apps.users.models import CustomUser
        from apps.orders.models import Order
        from django.db.models import Count, Sum, Avg
        
        customers = CustomUser.objects.filter(
            date_joined__date__gte=start_date,
            date_joined__date__lte=end_date,
            is_active=True
        )
        
        report_data = {
            'period': f"{start_date} to {end_date}",
            'new_customers': customers.count(),
            'total_active': CustomUser.objects.filter(is_active=True).count(),
        }
        
        # Customer value
        customer_stats = customers.annotate(
            order_count=Count('order'),
            total_spent=Sum('order__total_amount'),
            avg_order=Avg('order__total_amount')
        )
        
        report_data['customer_breakdown'] = [
            {
                'email': c.email,
                'orders': c.order_count,
                'total_spent': float(c.total_spent or 0),
                'avg_order': float(c.avg_order or 0),
            }
            for c in customer_stats[:100]
        ]
        
        return report_data
    
    @staticmethod
    def generate_inventory_report():
        """Generate inventory report"""
        from apps.products.models import Product
        from apps.inventory.models import InventorySKU
        from django.db.models import Sum
        
        products = Product.objects.filter(is_active=True)
        
        report_data = {
            'generated_at': timezone.now().isoformat(),
            'total_products': products.count(),
            'total_units': products.aggregate(Sum('stock'))['stock__sum'] or 0,
            'total_value': 0,
        }
        
        # Low stock items
        low_stock = products.filter(stock__lte=models.F('min_stock')).values_list('name', 'stock', 'sku')
        report_data['low_stock_items'] = list(low_stock)
        
        # Out of stock
        out_of_stock = products.filter(stock=0).count()
        report_data['out_of_stock_count'] = out_of_stock
        
        return report_data
    
    @staticmethod
    def export_to_excel(data, filename):
        """Export data to Excel"""
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        from django.core.files.base import ContentFile
        
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Write headers
        headers = list(data[0].keys()) if data else []
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Write data
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, (key, value) in enumerate(row_data.items(), 1):
                ws.cell(row=row_idx, column=col_idx).value = value
        
        # Save to file
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return ContentFile(output.read(), name=filename)
    
    @staticmethod
    def export_to_csv(data, filename):
        """Export data to CSV"""
        import csv
        from django.core.files.base import ContentFile
        import io
        
        output = io.StringIO()
        if data:
            writer = csv.DictWriter(output, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)
        
        return ContentFile(output.getvalue().encode('utf-8'), name=filename)
    
    @staticmethod
    def schedule_report(report):
        """Schedule report execution"""
        from celery import current_app
        from apps.reporting.tasks import execute_report
        
        # Calculate next run
        now = timezone.now()
        if report.schedule == 'daily':
            next_run = now + timedelta(days=1)
        elif report.schedule == 'weekly':
            next_run = now + timedelta(weeks=1)
        elif report.schedule == 'monthly':
            next_run = now + timedelta(days=30)
        
        report.next_run = next_run
        report.save()
        
        # Schedule task
        execute_report.apply_async(
            args=[report.id],
            eta=next_run
        )


# ============================================================
# CELERY TASKS
# ============================================================

"""
from celery import shared_task

@shared_task
def execute_scheduled_reports():
    '''Execute scheduled reports'''
    from apps.reporting.models import Report
    
    ready = Report.objects.filter(
        is_active=True,
        next_run__lte=timezone.now()
    )
    
    for report in ready:
        execute_report(report.id)

@shared_task
def execute_report(report_id):
    '''Execute single report'''
    from apps.reporting.models import Report, ReportExecution
    from django.core.mail import send_mail
    
    report = Report.objects.get(id=report_id)
    
    execution = ReportExecution.objects.create(
        report=report,
        start_time=timezone.now(),
        status='generating'
    )
    
    try:
        # Generate report
        if report.report_type == 'sales':
            data = ReportingEngine.generate_sales_report(
                timezone.now().date() - timedelta(days=30),
                timezone.now().date()
            )
        
        # Export
        if report.export_format == 'excel':
            file = ReportingEngine.export_to_excel(data, f'{report.name}.xlsx')
        else:
            file = ReportingEngine.export_to_csv(data, f'{report.name}.csv')
        
        execution.file = file
        execution.status = 'completed'
        
        # Email report
        for recipient in report.recipients:
            send_mail(
                f'Report: {report.name}',
                f'Your report is ready.',
                'reports@joankkfarm.com',
                [recipient],
            )
    
    except Exception as e:
        execution.status = 'failed'
        execution.error_message = str(e)
        logger.error(f'Report execution failed: {e}')
    
    execution.end_time = timezone.now()
    execution.save()

# Add to CELERY_BEAT_SCHEDULE:
'execute-scheduled-reports': {
    'task': 'apps.reporting.tasks.execute_scheduled_reports',
    'schedule': 3600.0,  # Hourly
},
"""